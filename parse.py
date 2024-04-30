from openpyxl import load_workbook
import json
wb = load_workbook(filename='中文图书馆目录.xlsx')
ws = wb.active
header = ws[1]
hvalues = []
for h in header:
    if h.value:
        hvalues.append(h.value)
    else:
        break
# print(hvalues)
pubindex = hvalues.index('Published At')
cntindex = hvalues.index('Page Count')
MAX_COL = len(hvalues)
MAX_ROW = 1436
with open('output.json', 'w') as f:
    for row in ws.iter_rows(min_row=2, max_row=MAX_ROW, min_col=1, max_col=MAX_COL, values_only=True):
        # print(row)
        rowlist = [str(x) if x else '' for x in row]
        rowlist[0] = rowlist[0].split('.')[0] # excel makes ID and ISBN as float, convert back to string
        rowlist[1] = rowlist[1].split('.')[0]
        rowlist[pubindex] = int(rowlist[pubindex].split('-')[0]) if rowlist[pubindex] else 0
        rowlist[cntindex] = int(rowlist[cntindex].split('.')[0]) if rowlist[cntindex] else 0
        d = dict(zip(hvalues, rowlist))
        # d['image_url'] = f'https://covers.openlibrary.org/b/isbn/{rowlist[1]}-M.jpg'
        if rowlist[1]:
            d['image_url'] = f'https://images.isbndb.com/covers/{rowlist[1][-4:-2]}/{rowlist[1][-2:]}/{rowlist[1]}.jpg'
        else:
            d['image_url'] = ''
        # print(d)
        json.dump(d, f, ensure_ascii=False)
        print(',', file=f)
