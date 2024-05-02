from openpyxl import load_workbook
import json,sys
wb = load_workbook(filename='中文图书馆目录.xlsx')
ws = wb.active
header = ws[1]
hvalues = []
for h in header:
    if h.value:
        hvalues.append(h.value)
    else:
        break
# print(hvalues)
pubindex = hvalues.index('Published At')
cntindex = hvalues.index('Page Count')
catindex = hvalues.index('Categories')
langindex = hvalues.index('Language')
publisherindex = hvalues.index('Publisher')
MAX_COL = len(hvalues)
MAX_ROW = 1436
labels = [x.lower().replace(' ', '_') for x in hvalues]
# print(labels)
labels[5]='name' #chant title to name
# print(labels)
# sys.exit(0)
books = []
with open('output2.json', 'w') as f:
    for row in ws.iter_rows(min_row=2, max_row=MAX_ROW, min_col=1, max_col=MAX_COL, values_only=True):
        # print(row)
        rowlist = [str(x).strip() if x else '' for x in row]
        rowlist[0] = rowlist[0].split('.')[0] # excel makes ID and ISBN as float, convert back to string
        rowlist[1] = rowlist[1].split('.')[0]
        rowlist[pubindex] = int(rowlist[pubindex].split('-')[0]) if rowlist[pubindex] else 0
        rowlist[cntindex] = int(rowlist[cntindex].split('.')[0]) if rowlist[cntindex] else 0
        rowlist[catindex] = rowlist[catindex].split(',') if rowlist[catindex] else ['未分类']
        rowlist[langindex] = '中文' if rowlist[langindex] != 'en' else '英文'
        rowlist[publisherindex] = rowlist[publisherindex] if rowlist[publisherindex] else '未知出版社'
        d = dict(zip(labels, rowlist))
        # d['image_url'] = f'https://covers.openlibrary.org/b/isbn/{rowlist[1]}-M.jpg'
        if rowlist[1]:
            d['image'] = f'https://images.isbndb.com/covers/{rowlist[1][-4:-2]}/{rowlist[1][-2:]}/{rowlist[1]}.jpg'
        else:
            d['image'] = 'https://dashboard.algolia.com/images/proxy/lT7Fbh5Y_u44AO9YxbYH63YM4QlmA1l4F-TVT8TCE34/resizing_type:fit/width:208/height:208/gravity:sm/enlarge:true/extend:true/aHR0cHM6Ly9pbWFnZXMuaXNibmRiLmNvbS9jb3ZlcnMvNDgvMjIvOTc4NzUxMzcxNDgyMi5qcGc.jpg'
        # print(d)
        books.append(d)
    json.dump(books, f, ensure_ascii=False, indent=2)
    # print(',', file=f)
